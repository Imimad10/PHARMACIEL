import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STATUT_FILL_HEX = {
    "Réglé": "D1FAE5",
    "En attente": "FEF3C7",
    "Refusée": "FEE2E2",
}

COLS = [
    ("N", "N°", 6),
    ("Chauffeur", "Chauffeur", 20),
    ("Client", "Client", 20),
    ("Montant", "Montant (DA)", 15),
    ("N_Cheque", "N° Chèque", 15),
    ("Date_Sortie", "Date de sortie", 15),
    ("Date_Retour", "Date de retour", 15),
    ("Statut", "Statut", 14),
]


def generate_cheques_excel(df, subtitle=""):
    """
    Génère un classeur Excel (.xlsx) stylé pour la liste des chèques.
    df doit contenir les colonnes : N, Chauffeur, Client, Montant, N_Cheque,
    Date_Sortie, Date_Retour, Statut.
    Retourne les bytes du fichier, prêts pour st.download_button.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi Chèques"

    n_cols = len(COLS)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Titre ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value="DarPharm Solution — Suivi des Chèques")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="364FC7", end_color="364FC7", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
        sub_cell.alignment = Alignment(horizontal="center")
        header_row = 4
    else:
        header_row = 3

    # --- En-têtes de colonnes ---
    for idx, (_, label, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        cell.fill = PatternFill(start_color="E8EBFF", end_color="E8EBFF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[header_row].height = 20

    # --- Données ---
    r = header_row + 1
    for _, row in df.iterrows():
        statut = str(row.get("Statut", ""))
        fill_hex = STATUT_FILL_HEX.get(statut, "FFFFFF")
        fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

        for idx, (key, _, _) in enumerate(COLS, start=1):
            val = row.get(key, "")
            if key == "Montant":
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = 0
            cell = ws.cell(row=r, column=idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if key == "Montant":
                cell.number_format = '#,##0 "DA"'
            if key == "Chauffeur" or key == "Client":
                cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    last_data_row = r - 1

    # --- Ligne de total ---
    if last_data_row >= header_row + 1:
        total_row = r
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        montant_col_idx = [i for i, (k, _, _) in enumerate(COLS, start=1) if k == "Montant"][0]
        col_letter = get_column_letter(montant_col_idx)
        total_cell = ws.cell(
            row=total_row,
            column=montant_col_idx,
            value=f"=SUM({col_letter}{header_row + 1}:{col_letter}{last_data_row})",
        )
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0 "DA"'
        total_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        for c in range(1, n_cols + 1):
            ws.cell(row=total_row, column=c).border = border
            if ws.cell(row=total_row, column=c).fill.start_color.rgb in (None, "00000000"):
                ws.cell(row=total_row, column=c).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # --- Filtre automatique + volet figé ---
    last_col_letter = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_data_row if last_data_row >= header_row else header_row}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()